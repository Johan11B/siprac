#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de normalización de datos meteorológicos
Procesa datos crudos de estaciones meteorológicas y los normaliza
"""

import sys
import json
import math
import unicodedata
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def normalizar_nombre_columna(nombre):
    nombre = str(nombre).strip().lower()
    nombre = unicodedata.normalize('NFKD', nombre)
    nombre = ''.join([c for c in nombre if not unicodedata.combining(c)])
    nombre = nombre.replace('%', '').replace('(', '').replace(')', '').replace('.', '').replace(':', '')
    nombre = nombre.replace('/', ' ').replace('-', ' ').replace('_', ' ').replace('\n', ' ').replace('\r', ' ')
    nombre = ' '.join(nombre.split())
    return nombre

def mapear_columna(nombre):
    nombre = normalizar_nombre_columna(nombre)

    if 'fecha' in nombre and 'hora' in nombre:
        return 'fecha_lectura'
    if 'timestamp' in nombre:
        return 'fecha_lectura'
    if 'intervalo' in nombre:
        return 'intervalo'
    if ('temp' in nombre or 'temperatura' in nombre) and ('interna' in nombre or 'inte' in nombre or ' i' in nombre):
        return 'temp_interna'
    if ('temp' in nombre or 'temperatura' in nombre) and ('externa' in nombre or 'exterior' in nombre or 'ext' in nombre or ' e' in nombre):
        return 'temp_externa'
    if 'humedad' in nombre and ('interna' in nombre or 'inte' in nombre or ' i' in nombre):
        return 'humedad_interna'
    if 'humedad' in nombre and ('externa' in nombre or 'exterior' in nombre or 'ex' in nombre or ' e' in nombre):
        return 'humedad_externa'
    if 'presion' in nombre and ('relati' in nombre or 'relativa' in nombre):
        return 'presion_relativa'
    if 'presion' in nombre and ('absolu' in nombre or 'absoluta' in nombre):
        return 'presion_absoluta'
    if 'velocidad' in nombre or ('vel' in nombre and 'viento' not in nombre and 'velocidad' not in nombre):
        return 'vel_viento'
    if 'viento' in nombre and 'vel' in nombre:
        return 'vel_viento'
    if 'rafaga' in nombre or 'rafaga' in nombre or 'rafága' in nombre:
        return 'rafaga'
    if 'direccion' in nombre:
        return 'direccion_viento'
    if 'punto' in nombre and 'roci' in nombre:
        return 'punto_rocio'
    if 'sensacion' in nombre or 'sensación' in nombre:
        return 'sensacion_termica'
    if 'lluvia' in nombre and '24' in nombre:
        return 'lluvia_24h'
    if 'lluvia' in nombre and 'hora' in nombre:
        return 'lluvia_hora'
    if 'lluvia' in nombre and 'semana' in nombre:
        return 'lluvia_semana'
    if 'lluvia' in nombre and 'mes' in nombre:
        return 'lluvia_mes'
    if 'lluvia' in nombre and 'total' in nombre:
        return 'lluvia_total'
    if 'lluvia' in nombre and 'm' in nombre and 'hora' not in nombre and '24' not in nombre and 'semana' not in nombre and 'mes' not in nombre and 'total' not in nombre:
        return 'lluvia'
    if nombre in ['n', 'no', 'numero', 'nº']:
        return 'numero'
    return None

def calcular_punto_rocio(temperatura, humedad):
    """
    Calcula el punto de rocío usando la fórmula de Magnus aproximada
    """
    if pd.isna(temperatura) or pd.isna(humedad) or humedad == 0:
        return np.nan
    
    a = 17.27
    b = 237.7
    alpha = ((a * temperatura) / (b + temperatura)) + np.log(humedad / 100.0)
    punto_rocio = (b * alpha) / (a - alpha)
    return round(punto_rocio, 1)

def calcular_sensacion_termica(temperatura, velocidad_viento, humedad):
    """
    Calcula la sensación térmica (Wind Chill si T < 10°C, o Heat Index si T > 26°C)
    """
    if pd.isna(temperatura) or pd.isna(velocidad_viento):
        return np.nan
    
    # Wind Chill (para temperaturas bajas)
    if temperatura < 10:
        wc = 13.12 + (0.6215 * temperatura) - (11.37 * (velocidad_viento ** 0.16)) + (0.3965 * temperatura * (velocidad_viento ** 0.16))
        return round(wc, 1)
    
    # Heat Index (para temperaturas altas)
    elif temperatura > 26:
        if not pd.isna(humedad):
            c1 = -42.379
            c2 = 2.04901523
            c3 = 10.14333127
            c4 = -0.22475541
            c5 = -0.00683783
            c6 = -0.05481717
            c7 = 0.00122874
            c8 = 0.00085282
            c9 = -0.00000199
            
            T = temperatura
            RH = humedad
            
            hi = (c1 + c2*T + c3*RH + c4*T*RH + c5*T**2 + c6*RH**2 + 
                  c7*T**2*RH + c8*T*RH**2 + c9*T**2*RH**2)
            return round(hi, 1)
    
    return temperatura

def normalizar_datos(input_file, output_file):
    """
    Normaliza los datos del archivo de entrada y guarda en archivo de salida
    """
    try:
        # Leer el archivo (intentar múltiples formatos)
        try:
            df = pd.read_excel(input_file)
        except:
            df = pd.read_csv(input_file)
        
        # Hacer una copia
        df = df.copy()
        
        # Normalizar y mapear columnas
        columnas_map = {}
        for col in df.columns:
            mapped = mapear_columna(col)
            if mapped:
                columnas_map[col] = mapped
        
        df = df.rename(columns=columnas_map)
        
        # Eliminar columnas de número si existen
        if 'numero' in df.columns:
            df = df.drop(columns=['numero'])

        # Asegurar que tenemos las columnas necesarias
        columnas_requeridas = ['fecha_lectura', 'temp_externa', 'humedad_externa', 'vel_viento']
        columnas_existentes = [col for col in columnas_requeridas if col in df.columns]

        if len(columnas_existentes) < 3:
            raise ValueError("El archivo no contiene suficientes columnas esperadas. Asegúrate de incluir Fecha/Hora, Temperatura Externa, Humedad Externa y Velocidad del Viento.")
        
        # Convertir fechas
        if 'fecha_lectura' in df.columns:
            df['fecha_lectura'] = pd.to_datetime(df['fecha_lectura'], errors='coerce')
        
        # Convertir columnas numéricas
        columnas_numericas = [
            'temp_interna', 'humedad_interna', 'temp_externa', 'humedad_externa',
            'presion_relativa', 'presion_absoluta', 'vel_viento', 'rafaga',
            'lluvia_hora', 'lluvia_24h', 'lluvia_semana', 'lluvia_mes', 'lluvia_total'
        ]

        for col in columnas_numericas:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        # Calcular punto de rocío
        if 'temp_externa' in df.columns and 'humedad_externa' in df.columns:
            df['punto_rocio'] = df.apply(
                lambda row: calcular_punto_rocio(row['temp_externa'], row['humedad_externa']),
                axis=1
            )
        
        # Calcular sensación térmica
        if 'temp_externa' in df.columns and 'vel_viento' in df.columns:
            humedad = df['humedad_externa'] if 'humedad_externa' in df.columns else None
            df['sensacion_termica'] = df.apply(
                lambda row: calcular_sensacion_termica(
                    row['temp_externa'],
                    row['vel_viento'],
                    row['humedad_externa'] if 'humedad_externa' in df.columns else None
                ),
                axis=1
            )
        
        # Agregar número de fila
        df.insert(0, 'N°', range(1, len(df) + 1))
        
        # Reordenar columnas de forma lógica
        columnas_orden = [
            'N°', 'fecha_lectura', 'intervalo', 'temp_interna', 'humedad_interna',
            'temp_externa', 'humedad_externa', 'presion_relativa', 'presion_absoluta',
            'vel_viento', 'rafaga', 'direccion_viento', 'punto_rocio', 'sensacion_termica',
            'lluvia_hora', 'lluvia_24h', 'lluvia_semana', 'lluvia_mes', 'lluvia_total'
        ]

        # Mantener solo columnas que existen
        columnas_finales = [col for col in columnas_orden if col in df.columns]
        df = df[columnas_finales]

        # Calcular estadísticas de resumen
        lluvia_total_valor = 0
        if 'lluvia_total' in df.columns:
            lluvia_total_valor = df['lluvia_total'].sum(skipna=True)
        elif 'lluvia_24h' in df.columns:
            lluvia_total_valor = df['lluvia_24h'].sum(skipna=True)
        elif 'lluvia_hora' in df.columns:
            lluvia_total_valor = df['lluvia_hora'].sum(skipna=True)

        summary = {
            'total_registros': int(len(df)),
            'fecha_inicio': df['fecha_lectura'].min().isoformat() if 'fecha_lectura' in df.columns else 'N/A',
            'fecha_fin': df['fecha_lectura'].max().isoformat() if 'fecha_lectura' in df.columns else 'N/A',
            'temp_promedio': float(round(df['temp_externa'].mean(), 2)) if 'temp_externa' in df.columns else 0.0,
            'temp_minima': float(round(df['temp_externa'].min(), 2)) if 'temp_externa' in df.columns else 0.0,
            'temp_maxima': float(round(df['temp_externa'].max(), 2)) if 'temp_externa' in df.columns else 0.0,
            'humedad_promedio': float(round(df['humedad_externa'].mean(), 1)) if 'humedad_externa' in df.columns else 0.0,
            'lluvia_total': float(round(lluvia_total_valor, 2)),
            'vel_viento_promedio': float(round(df['vel_viento'].mean(), 2)) if 'vel_viento' in df.columns else 0.0,
        }

        last_record = {}
        if len(df) > 0:
            last = df.iloc[-1]
            last_record = {
                'fecha_lectura': last['fecha_lectura'].isoformat() if 'fecha_lectura' in df.columns and pd.notna(last['fecha_lectura']) else None,
                'temp_externa': float(last['temp_externa']) if 'temp_externa' in df.columns and pd.notna(last['temp_externa']) else None,
                'humedad_externa': float(last['humedad_externa']) if 'humedad_externa' in df.columns and pd.notna(last['humedad_externa']) else None,
                'vel_viento': float(last['vel_viento']) if 'vel_viento' in df.columns and pd.notna(last['vel_viento']) else None,
                'lluvia_hora': float(last['lluvia_hora']) if 'lluvia_hora' in df.columns and pd.notna(last['lluvia_hora']) else None,
                'lluvia_24h': float(last['lluvia_24h']) if 'lluvia_24h' in df.columns and pd.notna(last['lluvia_24h']) else None,
                'lluvia_semana': float(last['lluvia_semana']) if 'lluvia_semana' in df.columns and pd.notna(last['lluvia_semana']) else None,
                'lluvia_mes': float(last['lluvia_mes']) if 'lluvia_mes' in df.columns and pd.notna(last['lluvia_mes']) else None,
                'lluvia_total': float(last['lluvia_total']) if 'lluvia_total' in df.columns and pd.notna(last['lluvia_total']) else None,
                'direccion_viento': str(last['direccion_viento']) if 'direccion_viento' in df.columns and pd.notna(last['direccion_viento']) else None,
                'punto_rocio': float(last['punto_rocio']) if 'punto_rocio' in df.columns and pd.notna(last['punto_rocio']) else None,
                'sensacion_termica': float(last['sensacion_termica']) if 'sensacion_termica' in df.columns and pd.notna(last['sensacion_termica']) else None,
            }

        # Asegurar tipos nativos para JSON
        for key, value in summary.items():
            if hasattr(value, 'item'):
                summary[key] = value.item()

        wb = Workbook()
        ws = wb.active
        ws.title = "Datos Normalizados"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Escribir encabezados
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Escribir datos
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                if col_idx == 1:  # N°
                    cell.alignment = Alignment(horizontal='center')
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Crear hoja de resumen
        ws_summary = wb.create_sheet("Resumen")
        ws_summary['A1'] = "RESUMEN DE NORMALIZACIÓN"
        ws_summary['A1'].font = Font(bold=True, size=12)
        
        summary_data = [
            ['', ''],
            ['Total de Registros', summary['total_registros']],
            ['Fecha de Inicio', summary['fecha_inicio']],
            ['Fecha de Fin', summary['fecha_fin']],
            ['', ''],
            ['Temperatura Promedio (°C)', summary['temp_promedio']],
            ['Temperatura Mínima (°C)', summary['temp_minima']],
            ['Temperatura Máxima (°C)', summary['temp_maxima']],
            ['', ''],
            ['Humedad Promedio (%)', summary['humedad_promedio']],
            ['Lluvia Total (m)', summary['lluvia_total']],
            ['Velocidad del Viento Promedio (m/s)', summary['vel_viento_promedio']],
        ]
        
        for row_idx, row_data in enumerate(summary_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                if col_idx == 1 and value:
                    cell.font = Font(bold=True)
        
        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 20
        
        # Guardar archivo
        wb.save(output_file)
        
        # Retornar resumen y último registro como JSON
        print(json.dumps({
            'summary': summary,
            'last_record': last_record,
        }))
        
    except Exception as e:
        print(json.dumps({'error': str(e)}), file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(json.dumps({'error': 'Argumentos insuficientes'}), file=sys.stderr)
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    normalizar_datos(input_file, output_file)
